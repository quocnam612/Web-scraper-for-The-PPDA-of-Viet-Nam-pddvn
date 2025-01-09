import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import os
import urllib.parse
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess
from datetime import datetime

BASE_URL = "https://ppdvn.gov.vn/web/guest/ke-hoach-xuat-ban"
HEADER = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
}

def get_total_pages(query):
    query_encoded = urllib.parse.quote(query)
    current_date = datetime.now().strftime("%d/%m/%Y")
    url = f"{BASE_URL}?query={query_encoded}&id_nxb=8&bat_dau=01%2F01%2F2004&ket_thuc={current_date}&p=1"

    try:
        response = requests.get(url, headers=HEADER)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")
        pagination = soup.find("div", {"class": "pagination"})

        if not pagination:
            print("No pagination found")
            return 1

        # Find all <a> tags inside pagination and extract page numbers from "href"
        links = pagination.find_all("a")
        page_numbers = []

        for link in links:
            href = link.get("href")
            if href and "p=" in href:  # Look for page numbers in "p="
                page_param = href.split("p=")[-1]  # Get the value after "p="
                if page_param.isdigit():  # Ensure it's numeric
                    page_numbers.append(int(page_param))

        if not page_numbers:
            print("No valid page numbers found")
            return 1

        # Return the highest page number
        total_pages = max(page_numbers)
        return total_pages

    except Exception as e:
        print(f"Error occurred: {e}")
        return 1

def get_figures(key, query):
    query_encoded = urllib.parse.quote(query)
    current_date = datetime.now().strftime("%d/%m/%Y")
    url = f"{BASE_URL}?query={query_encoded}&id_nxb=8&bat_dau=01%2F01%2F2004&ket_thuc={current_date}&p={key}"
    try:
        response = requests.get(url, headers=HEADER)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("table", {"cellpadding": "0", "cellspacing": "0"})
        
        if not table:
            print("No data found")
            return None
        
        tbody = table.find("tbody")
        if not tbody:
            print("No tbody found in the table")
            return None
        
        rows = tbody.find_all("tr")
        if not rows:
            print("No rows found in the tbody")
            return None
        
        data = []
        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 8:
                print("Not enough cells in the row")
                continue

            row_data = {
                "STT": cells[0].text.strip(),
                "ISBN": cells[1].text.strip(),
                "Title": cells[2].text.strip(),
                "Author": cells[3].text.strip(),
                "Translator": cells[4].text.strip(),
                "Numbers of ordered copies": cells[5].text.strip(),
                "Self-publishing": cells[6].text.strip(),
                "Partner": cells[7].text.strip()
            }
            data.append(row_data)
        
        return data
    
    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

def scrape_figures(title, startpage, endpage, output, sheet_name):
    all_data = []
    for key in range(startpage, endpage + 1):
        print(f"Fetching data for page number {key}")
        page_data = get_figures(key, title)
        if page_data:
            all_data.extend(page_data)
    
    if all_data:
        df = pd.DataFrame(all_data)
        if os.path.exists(output):
            book = load_workbook(output)
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            with pd.ExcelWriter(output, engine='openpyxl', mode='a') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Data saved to {output} in sheet {sheet_name}")
        subprocess.Popen(["start", output], shell=True)
    else:
        print(f"No data found for the title '{title}'")

def start_scraping():
    title = url_entry.get()
    output = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    sheet_name = title

    if not title or not output:
        messagebox.showerror("Input Error", "Please provide all required inputs.")
        return

    total_pages = get_total_pages(title)
    scrape_figures(title, 1, total_pages, output, sheet_name)
    messagebox.showinfo("Success", f"Data scraping completed. Data saved to {output}")

root = tk.Tk()
root.title("Web Scraper App")

# URL Input
tk.Label(root, text="Enter Title:").grid(row=0, column=0, padx=10, pady=10)
url_entry = tk.Entry(root, width=50)
url_entry.grid(row=0, column=1, padx=10, pady=10)

# Buttons
scrape_button = tk.Button(root, text="Start Scraping", command=start_scraping)
scrape_button.grid(row=1, column=0, columnspan=2, pady=10)

# Run the application
root.mainloop()
